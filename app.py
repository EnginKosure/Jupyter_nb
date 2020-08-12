# A library to handle excel workbooks
import openpyxl as xl
from openpyxl.chart import BarChart, Reference
# Load a workbook
wb = xl.load_workbook('transactions.xlsx')
# print(wb['Sheet1'])
# Get a Sheet from that workbook
sheet = wb['Sheet1']
# print(sheet)
# This is how to reach to a cell
cell = sheet['A1']
# You can also use (row, col) notation
# But first row is 1, first column is 1, not ZERO.
# cell = sheet.cell(1, 1)
# We can see only the memory reference for now.To see inside we will loop through it.
print(cell)  # <Cell 'Sheet1'.A1>
# This shows the highest row number
print(sheet.max_row)
# Cell
for row in range(2, sheet.max_row+1):
    cell = sheet.cell(row, 3)
    # To manipulate inside of a cell, we must reach it with .value
    corrected_price = cell.value*0.9
    print(corrected_price)
    # Get the adjacent column, cell by cell, in each iteration
    corrected_new_val = sheet.cell(row, 4)
    # Assign the value to the newly generated cells
    corrected_new_val.value = corrected_price

values = Reference(sheet, min_row=2, max_row=sheet.max_row,
                   min_col=4, max_col=4)
# make a chart
chart = BarChart()
# Add data to chart
chart.add_data(values)
# Add chart to the cell e2
sheet.add_chart(chart, 'e2')
# Save the workbook with a different name
wb.save('transactions2.xlsx')
